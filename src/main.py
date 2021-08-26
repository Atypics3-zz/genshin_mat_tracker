from openpyxl import load_workbook
import re

def main():
    wb = load_workbook(filename="src/mat_requirements.xlsx")
    sheet = wb['Requirement Tables']

    for x in range(1,9):
        print(x, sheet.cell(column=x, row=x).value)
    


if __name__ == '__main__':
    main()